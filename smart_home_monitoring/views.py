from django.http import HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth import authenticate, login, logout
from django.views.generic import TemplateView, View
from django.contrib.auth.models import User
from django.urls import reverse_lazy
from openpyxl import load_workbook
import pandas as pd
import os
from keras.models import load_model
from sklearn.preprocessing import StandardScaler

from smart_home_monitoring.forms import UploadDataForm
from .converters import ResultSummary, OrderedLabelEncoder, TextLabelEncoderDummy





def index(request):
    # template_name = "index.html"
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        print("\nUser Name = ",username)
        print("Password = ",password)
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('smart_home_monitoring:dashboard')
            
        else:
            context = {'message':'Invalid User Name and Password'}
            return render(request, 'index.html', context)
    return render(request, 'index.html', {'name': 'admin', 'pass': 'Info@123'})

        

def logoutUser(request):
    logout(request)
    return redirect('home')


class DashboardView(LoginRequiredMixin,TemplateView):
    template_name = "dashboard.html"

    def get_context_data(self, **kwargs):
        context = super(DashboardView, self).get_context_data(**kwargs)
        # context['cls'] = Lecturer.objects.count()
        # context['results'] = Results.objects.count()
        # context['students'] = Student.objects.count()
        # context['subjects'] = Subject.objects.count()
        return context
    

# Build paths inside the project like this: BASE_DIR / 'subdir'.
# BASE_DIR = Path(__file__).resolve().parent.parent
directory = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))    

model = os.path.join(directory, 'smart_home_monitoring/smart_home/smart_home_model.h5')
# print(model)
activity = pd.read_csv(os.path.join(directory, 'smart_home_monitoring/smart_home/activity.csv'), index_col=False) 


def upload_data(request):
    context = {}
    if request.method == "POST":
        form = UploadDataForm(request.POST, request.FILES)
        print(request.FILES.get('excel_file'))
        if form.is_valid():
            file = request.FILES.get('excel_file')
            pred_range = request.POST['pred_range']
        
            wb = load_workbook(file, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            df = pd.DataFrame(sheet.values)

            df.columns = df.iloc[0]
            yy_train_df = activity #df['activity']
            yy = df['activity']
            df = df.iloc[1:,]
            print(yy_train_df)

            X_test_df = df.drop(labels=['numDistinctSensors','sensorCount-Chair','sensorCount-DiningRoom',
             'sensorCount-Hall','sensorCount-Office','sensorCount-WorkArea','sensorElTime-Chair','sensorElTime-DiningRoom',
             'sensorElTime-Hall','sensorElTime-Office',
             'sensorElTime-WorkArea','lastSensorLocation','lastMotionLocation', 'sensorElTime-OutsideDoor', 'lastSensorEventSeconds','activity'], axis=1)

            X_test_df = X_test_df.to_numpy() # test
            # Scaling data
            sc = StandardScaler()
            # ss_train = sc.fit_transform(xx_train_df) # train
            S_test = sc.fit_transform(X_test_df)
            # s_val_pred = sc.transform(ss_val_df)
            # print(S_test)

            encoded_Y_pred, encoder = TextLabelEncoderDummy.labelencoder(yy_train_df)
            dummy_y_pred, uniques = TextLabelEncoderDummy.encoded_to_dummy(encoded_Y_pred)
            print('uniques '+str(uniques))
            saved_model = load_model(model)
            # saved_model.summary()

            # 5. make predictions
            pred = saved_model.predict(S_test, verbose=1)
            print(pred.shape, uniques.shape)

            # Reverse predictios
            reverse_dummy_pred =TextLabelEncoderDummy.reverse_dummy_to_encoded(pred,uniques)
            print(pd.unique(reverse_dummy_pred).shape)

            reverse_encoded_y_pred = TextLabelEncoderDummy.reverse_encoded_to_text(reverse_dummy_pred,encoder)
            print(reverse_encoded_y_pred.shape)
            yy=yy.iloc[1:,]
            object_list = zip(reverse_encoded_y_pred,yy)
            print(yy.shape)

            field_list = ['Actual', 'Predicted']
                
            context={'panel_title':'Smart Home Simulator',
            'field_list':field_list,
            'object_list':object_list}
            return render(request, "smart_home/smarthome_list.html", context)

        else:
            print(form.errors)
            # messages.success(request, 'Account was created for ' + username)
            return redirect('smart_home_monitoring:upload_data')
    else:
        form = UploadDataForm()
        context['main_page_title'] = 'Declare Students Result'
        context['panel_name'] = 'Results'
        context['panel_title'] = 'Declare Result'
        context['form'] = form
    return render(request, "smart_home/smarthome_list.html", context)








