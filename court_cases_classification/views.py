from django.http import HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth import authenticate, login, logout
from django.views.generic import TemplateView, View
from django.contrib.auth.models import User
from django.urls import reverse_lazy
from openpyxl import load_workbook
import os
import itertools

# from sklearn.preprocessing import StandardScaler

from court_cases_classification.forms import UploadCourtCaseForm
from .court_cases.predict import Predict
import pandas as pd





def logoutUser(request):
    logout(request)
    return redirect('home')


class DashboardView(LoginRequiredMixin,TemplateView):
    template_name = "dashboard.html"

    def get_context_data(self, **kwargs):
        context = super(DashboardView, self).get_context_data(**kwargs)
        # context['cls'] = Lecturer.objects.count()
        # context['results'] = Results.objects.count()
        # context['students'] = Student.objects.count()
        # context['subjects'] = Subject.objects.count()
        return context
    

def index(request):
    # template_name = "index.html"
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        print("\nUser Name = ",username)
        print("Password = ",password)
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('court_cases_classification:dashboard')
            
        else:
            context = {'message':'Invalid User Name and Password'}
            return render(request, 'index.html', context)
    return render(request, 'index.html', {'name': 'smart', 'pass': 'Home@123'})

directory = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))    
file_db = os.path.join(directory, 'court_cases_classification/court_cases/sherloc_court_cases_7.csv')

# def save_record(df3,file_db):
#     # save record
#     df3.to_csv(file_db,index=False)


def upload_data(request):
    context = {}
    if request.method == "POST":
        form = UploadCourtCaseForm(request.POST, request.FILES)
        print(request.FILES.get('excel_file'))
        if form.is_valid():
            file = request.FILES.get('excel_file')
            text = request.POST['text']
            if file != None:
                # this will take from excel sheets
                wb = load_workbook(file, data_only=True)
                sheet = wb[wb.sheetnames[0]]
                df = pd.DataFrame(sheet.values,columns=['text'])
                # print(df)
                data = itertools.chain.from_iterable(df.values)
                # print(data)
            else:
                # transform data
                # text = pred_text.get('1.0','end')
                df = pd.DataFrame({'text':[text]})
                data = df.values
            X = Predict.transform_text(df)
            predR = Predict.pred_r(X)
            predC = Predict.pred_c(X)

            predRList = []
            for i in predR:
                if i == 0:
                    predRList.append('Not Sentence')
                else:
                    predRList.append('Sentenced')
            # data = [text]
            # print(predRList)
            # print(predC)
            # save record
            # df1 = pd.DataFrame({'text':data,'crime_types':predC,'sentence':predRList})
            # df2 = pd.read_csv(file_db)
            # df3 = pd.concat([df2,df1],axis=0)
            # df3.to_csv(file_db,index=False)


            object_list = zip(data,predRList,predC)
            field_list = ['Court Case(s)','Sentence Status', 'Crime Type']
            context={'panel_title':'Court Cases Model Simulator',
            'field_list':field_list,
            'object_list':object_list}

            return render(request, "court_cases/court_cases_list.html", context)

        else:
            print(form.errors)
            # messages.success(request, 'Account was created for ' + username)
            return redirect('court_cases_classification:court_case')
    else:
        form = UploadCourtCaseForm()
        context['main_page_title'] = 'Declare Students Result'
        context['panel_name'] = 'Results'
        context['panel_title'] = 'Declare Result'
        context['form'] = form
        return render(request, "court_cases/court_cases_list.html", context)







